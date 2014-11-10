from __future__ import print_function
from openpyxl import load_workbook
import xlsxwriter
import sys
import string

#Create file from user input
if len(sys.argv) < 2:
    sys.exit("Error: need filename as input")

fileName = sys.argv[1] 
file = open(fileName)
fileNameWithoutExt = fileName.split(".")
newFile = fileNameWithoutExt[0] + ".xlsx"
readfilename = "3980edges_V.xlsx"
coordFile = open("coordFile.txt", 'a')
k = int(sys.argv[2])

edges = {}
values = []

for line in file:
    entries = string.split(line)
    key = int(entries[0])
    value = int(entries[1])

    if key in edges:
        edges[key].append(value)
    else:
        edges[key] = [value]
    if not value in values:
        values.append(value)

keys = sorted(edges.keys())
values = sorted(values)
for value in values:
    if not value in keys:
        keys.append(value)
keys = sorted(keys)


#Create a workbook and add a worksheet
workbook = xlsxwriter.Workbook(newFile);
worksheet = workbook.add_worksheet()

#Fill headings of adjacency matrix
LUT = {}
for i in range(1, len(keys) + 1):
    worksheet.write(i, 0, keys[i-1])
    worksheet.write(0, i, keys[i-1])
    value = keys[i-1]
    LUT[i] = value

#print(LUT)
#print(edges)

#Fill adjacency matrix
for col in range(1, len(keys) + 1):
    key = LUT[col]
    for row in range(1, len(keys) + 1):
        value = LUT[row]
        if key in edges:
            if value in edges[key]:
                worksheet.write(row, col, 1)
            else:
                worksheet.write(row, col, 0)
        else:
            worksheet.write(row, col, 0)

workbook.close() 
print("Successfully written to disk")

#Read data from values 
"""wb = load_workbook(filename = readfilename)
ws_unnorm = wb.get_sheet_by_name(name = "Unnormalized")
ws_sym = wb.get_sheet_by_name(name = "Symmetric")
ws_rw = wb.get_sheet_by_name(name = "Random-walk")

row_num = 1;
for row in ws_unnorm.rows:
    col_num = 1
    coord_str = str(row_num) + " "
    for cell in row:
        if not col_num == 1:
            coord_str += str(cell.internal_value) + " "
            if col_num > k:
                coordFile.write(coord_str + "\n")
                break
        col_num += 1
    row_num += 1
        
coordFile.close() """
    
